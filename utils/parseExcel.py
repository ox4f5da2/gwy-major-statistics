from openpyxl import load_workbook
import re

def parseJ(column_J_data):
  if "仅限" in column_J_data:
    education = [column_J_data.split("仅限")[1]]
  elif "或" in column_J_data:
    education = column_J_data.split("或")
  elif "及以上" in column_J_data:
    if "高中" in column_J_data:
      education = ["高中(中专)", "本科", "硕士研究生", "博士研究生"]
    elif "专科" in column_J_data:
      education = ["专科", "本科", "硕士研究生", "博士研究生"]
    elif "本科" in column_J_data:
      education = ["本科", "硕士研究生", "博士研究生"]
    elif "硕士研究生" in column_J_data:
      education = ["硕士研究生", "博士研究生"]
    elif "博士研究生" in column_J_data:
      education = ["博士研究生"]
    else:
      education = []
  elif "博士研究生" in column_J_data:
    education = ["博士研究生"]
  else:
      education = []
  return education


def parseN(column_N_data):
  if "不限" in column_N_data:
    political_status = ["中共党员", "共青团员"]
  elif "或" in column_N_data:
    political_status = column_N_data.split("或")
  else:
    political_status = [column_N_data]
  return political_status


def parseM(column_M_data):
  return "高校毕业生" in column_M_data


def parseQ(column_Q_data):
  if "三级专业目录" in column_Q_data:
    matches = re.findall(r"要求为：(.*+)", column_Q_data)
    # 使用正则表达式匹配出需要的内容
    matches = re.findall(r"三级专业目录(.*?类)：不限|([^\s]+)", matches[0])
    result = []
    for (val1, val2) in matches:
      if val1 != '' and val1 != '或者':
        result.append(val1)
      if val2 != '' and val2 != '或者':
        result.append(val2)
    return result
  elif "不限" in column_Q_data:
    return ["不限"]
  else:
    match = re.search(r"：(.+)", column_Q_data)
    if match:
      # 获取冒号后的文字部分
      requirements_text = match.group(1)
      return [s.strip() for s in requirements_text.split("或者")]
    else:
      return []


def parse_excel(filename):
  # 打开 Excel 文件
  wb = load_workbook(filename)
  sheet_count = len(wb.sheetnames)
  
  # 存储解析后的数据的列表
  parsed_data = []

  for index in range(sheet_count):
    # 遍历每一行
    for row in wb.worksheets[index].iter_rows(min_row=2, values_only=True):
      # 解析第 Q 列，处理专业
      column_Q_data = row[16]
      major = parseQ(column_Q_data)

      # 解析第 J 列，统计学历是本科还是研究生
      column_J_data = row[9]
      education = parseJ(column_J_data)

      # 解析第 N 列，统计政治面貌是党员还是共青团员
      column_N_data = row[13]
      political_status = parseN(column_N_data)

      # 存储解析后的数据到列表中
      parsed_data.append({
        "专业": major,
        "学历": education,
        "政治面貌": political_status
      })

  return parsed_data